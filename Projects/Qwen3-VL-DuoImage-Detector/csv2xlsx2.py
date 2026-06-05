#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CSV -> Excel(xlsx) 转换脚本（中文编码友好） + 批量目录转换

新增：
- 支持目录批量转换（可选递归）
- 输出到指定 xlsx_output 目录，保持相对路径结构
- 保留单文件模式以及原有参数（encoding / delimiter / sheet-name / autowidth）

示例：
# 批量（不递归）
python csv2xlsx.py --input-dir ./csv --output-dir ./xlsx_output --autowidth

# 批量（递归）
python csv2xlsx.py -I ./csv -O ./xlsx_output -R --autowidth
"""

from __future__ import annotations
import argparse
import csv
from pathlib import Path
from typing import Optional, Iterable

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ------------------------ 工具函数 ------------------------

def detect_encoding(p: Path, user_encoding: Optional[str] = None) -> str:
    """尝试判断文件编码。优先使用用户指定；否则按常见编码依次尝试。"""
    if user_encoding and user_encoding.lower() != "auto":
        return user_encoding

    candidates = [
        "utf-8-sig",  # 带 BOM 的 UTF-8
        "utf-8",
        "gbk",
        "gb18030",
        "big5",
        "cp936",    # Windows 简体中文
    ]
    data = p.read_bytes()
    for enc in candidates:
        try:
            data.decode(enc)
            return enc
        except Exception:
            continue
    # 兜底
    return "utf-8"


def sniff_delimiter(sample_text: str, default: str = ",") -> str:
    try:
        sniffer = csv.Sniffer()
        dialect = sniffer.sniff(sample_text, delimiters=[",", "\t", ";", "|"])
        return dialect.delimiter
    except Exception:
        return default


def iter_csv_rows(path: Path, encoding: str, delimiter: Optional[str]) -> tuple[str, list[list[str]]]:
    # 读取样本用于嗅探分隔符
    with path.open("r", encoding=encoding, errors="strict") as f:
        head = f.read(4096)
    delim = delimiter if (delimiter and delimiter.lower() != "auto") else sniff_delimiter(head)

    rows: list[list[str]] = []
    with path.open("r", encoding=encoding, errors="strict", newline="") as f:
        rdr = csv.reader(f, delimiter=delim)
        for row in rdr:
            rows.append([cell for cell in row])
    return delim, rows


def write_to_xlsx(rows: list[list[str]], out_path: Path, sheet_name: str, auto_width: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    # openpyxl 限制：工作表名最长 31 字符，且不能含 \\ / * ? : [ ]
    safe_title = sheet_name[:31]
    for ch in "\\/*?:[]":
        safe_title = safe_title.replace(ch, " ")
    ws.title = safe_title

    # 写入单元格
    max_col_width = {}
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            v = val if val is not None else ""
            ws.cell(row=r_idx, column=c_idx, value=v)
            if auto_width:
                ln = len(str(v))
                max_col_width[c_idx] = max(max_col_width.get(c_idx, 0), ln)

    if auto_width and max_col_width:
        for c_idx, width in max_col_width.items():
            # 适当加一点余量，中文字符在 Excel 中宽度略大
            ws.column_dimensions[get_column_letter(c_idx)].width = min(80, max(10, width * 1.2))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))


# ------------------------ 批量相关 ------------------------

def find_csv_files(root: Path, recursive: bool = False) -> Iterable[Path]:
    if recursive:
        yield from (p for p in root.rglob("*") if p.is_file() and p.suffix.lower() == ".csv")
    else:
        yield from (p for p in root.glob("*.csv") if p.is_file())


def convert_single_csv(
    in_path: Path,
    out_path: Path,
    encoding_opt: str = "auto",
    delimiter_opt: str = "auto",
    sheet_name: str = "数据",
    autowidth: bool = False,
) -> None:
    enc = detect_encoding(in_path, encoding_opt)
    delim, rows = iter_csv_rows(in_path, enc, delimiter_opt)
    print(f"[INFO] {in_path.name} -> 编码: {enc} | 分隔符: {repr(delim)} | 行数: {len(rows)}")
    write_to_xlsx(rows, out_path, sheet_name, auto_width=autowidth)
    print(f"[OK] 已生成: {out_path}")


def batch_convert(
    in_dir: Path,
    out_dir: Path,
    recursive: bool,
    encoding_opt: str,
    delimiter_opt: str,
    sheet_name: str,
    autowidth: bool,
) -> None:
    if not in_dir.exists() or not in_dir.is_dir():
        raise SystemExit(f"输入目录不存在或不是目录：{in_dir}")

    cnt_total = 0
    cnt_ok = 0
    for csv_path in find_csv_files(in_dir, recursive=recursive):
        cnt_total += 1
        # 计算输出路径：保持相对结构（相对 in_dir），替换后缀为 .xlsx
        rel = csv_path.relative_to(in_dir)
        out_path = out_dir / rel.with_suffix(".xlsx")
        try:
            convert_single_csv(
                csv_path, out_path,
                encoding_opt=encoding_opt,
                delimiter_opt=delimiter_opt,
                sheet_name=sheet_name,
                autowidth=autowidth
            )
            cnt_ok += 1
        except Exception as e:
            print(f"[ERR] 转换失败：{csv_path} -> {e}")

    print(f"\n[SUMMARY] 目标目录：{in_dir}")
    print(f"总计发现 CSV：{cnt_total} | 成功：{cnt_ok} | 失败：{cnt_total - cnt_ok}")
    print(f"输出目录：{out_dir.resolve()}")


# ------------------------ 主流程 ------------------------

def main():
    ap = argparse.ArgumentParser(description="CSV 转 Excel (xlsx) —— 中文编码友好 & 支持目录批量")
    # 单文件模式
    ap.add_argument("-i", "--input", help="输入 CSV 文件路径（单文件模式）")
    ap.add_argument("-o", "--output", help="输出 XLSX 文件路径（单文件模式）")
    # 目录模式
    ap.add_argument("-I", "--input-dir", help="输入 CSV 目录路径（批量模式）")
    ap.add_argument("-O", "--output-dir", help="输出 XLSX 目录路径（批量模式，默认 <input-dir>/../xlsx_output）")
    ap.add_argument("-R", "--recursive", action="store_true", help="递归处理子目录（批量模式）")

    # 通用选项
    ap.add_argument("-e", "--encoding", default="auto", help="CSV 编码，默认 auto 自动识别")
    ap.add_argument("-d", "--delimiter", default="auto", help="分隔符，默认 auto 自动嗅探，可指定如 ',' 或 '\\t'")
    ap.add_argument("-s", "--sheet-name", default="数据", help="工作表名称，默认 '数据'")
    ap.add_argument("--autowidth", action="store_true", help="根据内容粗略自适应列宽")

    args = ap.parse_args()

    # 判定模式
    single_mode = bool(args.input) or bool(args.output)
    batch_mode = bool(args.input_dir) or bool(args.output_dir)

    if single_mode and batch_mode:
        raise SystemExit("单文件模式与批量模式参数不可混用，请只使用一种模式。")

    if batch_mode:
        in_dir = Path(args.input_dir) if args.input_dir else None
        if in_dir is None:
            raise SystemExit("批量模式需要提供 --input-dir")
        out_dir = Path(args.output_dir) if args.output_dir else (in_dir.parent / "xlsx_output")
        batch_convert(
            in_dir=in_dir,
            out_dir=out_dir,
            recursive=args.recursive,
            encoding_opt=args.encoding,
            delimiter_opt=args.delimiter,
            sheet_name=args.sheet_name,
            autowidth=args.autowidth,
        )
        return

    # 单文件模式（与原始脚本兼容）
    if not (args.input and args.output):
        raise SystemExit("单文件模式需要 -i/--input 与 -o/--output 同时提供；或改用批量模式 -I/--input-dir。")

    in_path = Path(args.input)
    out_path = Path(args.output)

    if not in_path.exists():
        raise SystemExit(f"输入文件不存在：{in_path}")

    convert_single_csv(
        in_path,
        out_path,
        encoding_opt=args.encoding,
        delimiter_opt=args.delimiter,
        sheet_name=args.sheet_name,
        autowidth=args.autowidth,
    )


if __name__ == "__main__":
    main()
